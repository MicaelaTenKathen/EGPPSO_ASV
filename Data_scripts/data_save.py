import openpyxl


def savexlsx(MSE_data, sigma_data, mu_data, distances, it, e1, e2, e3, e4, e5):
    for i in range(len(mu_data)):
        mu_data[i] = float(mu_data[i])

    wb1 = openpyxl.Workbook()
    hoja1 = wb1.active
    hoja1.append(MSE_data)
    wb1.save(e1)

    wb2 = openpyxl.Workbook()
    hoja2 = wb2.active
    hoja2.append(sigma_data)
    wb2.save(e2)

    wb3 = openpyxl.Workbook()
    hoja3 = wb3.active
    hoja3.append(mu_data)
    wb3.save(e3)

    wb4 = openpyxl.Workbook()
    hoja4 = wb4.active
    hoja4.append(list(distances))
    wb4.save(e4)

    wb5 = openpyxl.Workbook()
    hoja5 = wb5.active
    hoja5.append(it)
    wb5.save(e5)