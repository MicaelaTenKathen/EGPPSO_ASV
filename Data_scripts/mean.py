import openpyxl
import numpy as np


def data_val(n):
    wb2 = openpyxl.load_workbook(n)
    data = list()
    h = 0
    hoja2 = wb2.active
    while True :
        h += 1
        print(h)
        cel2 = hoja2.cell(row=1, column=h)
        l = cel2.value
        if l == None:
            del(data[-1])
            break
        else:
            data.append(cel2.value)
    return data

def error_m(multmin, multmax, b, u,
            b20, b3, b12, b123, b145, b147, b156, b158, b159, b258, b321, b357, b369, b456, b4875, b541, b57, b65, b654, b741, b753, b756, b789, b8462, b852, b89, b95, b951, b963, b987):
    wb3 = openpyxl.load_workbook(u)
    hoja3 = wb3.active
    array = globals()['data%s' % seed[i]]
    mse = list()
    if seed[i] == 20:
        b = b20
    elif seed[i] == 3:
        b = b3
    elif seed[i] == 12:
        b = b12
    elif seed[i] == 123:
        b = b123
    elif seed[i] == 145:
        b = b145
    elif seed[i] == 147:
        b = b147
    elif seed[i] == 156:
        b = b156
    elif seed[i] == 158:
        b = b158
    elif seed[i] == 159:
        b = b159
    elif seed[i] == 258:
        b = b258
    elif seed[i] == 321:
        b = b321
    elif seed[i] == 357:
        b = b357
    elif seed[i] == 369:
        b = b369
    elif seed[i] == 456:
        b = b456
    elif seed[i] == 4875:
        b = b4875
    elif seed[i] == 541:
        b = b541
    elif seed[i] == 57:
        b = b57
    elif seed[i] == 65:
        b = b65
    elif seed[i] == 654:
        b = b654
    elif seed[i] == 741:
        b = b741
    elif seed[i] == 753:
        b = b753
    elif seed[i] == 756:
        b = b756
    elif seed[i] == 789:
        b = b789
    elif seed[i] == 8462:
        b = b8462
    elif seed[i] == 852:
        b = b852
    elif seed[i] == 89:
        b = b89
    elif seed[i] == 95:
        b = b95
    elif seed[i] == 951:
        b = b951
    elif seed[i] == 963:
        b = b963
    elif seed[i] == 987:
        b = b987
    for j in range(len(array)):
        if multmin <= array[j] < multmax:
            b += 1
            print(b)
            cel3 = hoja3.cell(row=1, column=b)
            mse.append(cel3.value)
    if seed[i] == 20:
        b20 = b
    elif seed[i] == 3:
        b3 = b
    elif seed[i] == 12:
        b12 = b
    elif seed[i] == 123:
        b123 = b
    elif seed[i] == 145:
        b145 = b
    elif seed[i] == 147:
        b147 = b
    elif seed[i] == 156:
        b156 = b
    elif seed[i] == 158:
        b158 = b
    elif seed[i] == 159:
        b159 = b
    elif seed[i] == 258:
        b258 = b
    elif seed[i] == 321:
        b321 = b
    elif seed[i] == 357:
        b357 = b
    elif seed[i] == 369:
        b369 = b
    elif seed[i] == 456:
        b456 = b
    elif seed[i] == 4875:
        b4875 = b
    elif seed[i] == 541:
        b541 = b
    elif seed[i] == 57:
        b57 = b
    elif seed[i] == 65:
        b65 = b
    elif seed[i] == 654:
        b654 = b
    elif seed[i] == 741:
        b741 = b
    elif seed[i] == 753:
        b753 = b
    elif seed[i] == 756:
        b756 = b
    elif seed[i] == 789:
        b789 = b
    elif seed[i] == 8462:
        b8462 = b
    elif seed[i] == 852:
        b852 = b
    elif seed[i] == 89:
        b89 = b
    elif seed[i] == 95:
        b95 = b
    elif seed[i] == 951:
        b951 = b
    elif seed[i] == 963:
        b963 = b
    elif seed[i] == 987:
        b987 = b
    return mse, b20, b3, b12, b123, b145, b147, b156, b158, b159, b258, b321, b357, b369, b456, b4875, b541, b57, b65, b654, b741, b753, b756, b789, b8462, b852, b89, b95, b951, b963, b987


#seed = [20, 95, 541, 65, 145, 541, 156, 158, 12, 3, 89, 57, 123, 456, 789, 987, 654, 321, 147, 258, 369, 741, 852, 963, 159, 951, 753, 357, 756, 8462, 4875]
seed = [20, 3, 12, 57, 65, 89, 95, 123, 145, 147, 156, 158, 159, 258, 321, 357, 369, 456, 541, 654, 741, 753, 756, 789, 852, 951, 963, 987, 4875, 8462]
dist = list()
mse_mean = list()
mse_std = list()
it = list()

d = dict
for i in range(len(seed)):
    globals()['data%s' % seed[i]] = data_val(str(r'C:\Users\mcjara\OneDrive - Universidad Loyola Andalucía\Documentos\PycharmProjects\PSO_ASVs\Pruebas\Ori\Data'+str(seed[i])+'.xlsx'))
    print('in')

multmin = 0
mse400, mse800, mse1200, mse1600, mse2000, mse2400, mse2800, mse3200, mse3600, mse4000, mse4400, mse4800, mse5200, mse5600, mse6000 = list(), list(), list(), list(), list(), list(), list(), list(), list(), list(), list(), list(), list(), list(), list()
b20, b3, b12, b123, b145, b147, b156, b158, b159, b258, b321, b357, b369, b456, b4875, b541, b57, b65, b654, b741, b753, b756, b789, b8462, b852, b89, b95, b951, b963, b987 = 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0
mult = list()

while multmin < 6001:
    multmax = multmin + 400
    mult.append(multmin)
    for i in range(len(seed)):
        b = 0
        mse, b20, b3, b12, b123, b145, b147, b156, b158, b159, b258, b321, b357, b369, b456, b4875, b541, b57, b65, b654, b741, b753, b756, b789, b8462, b852, b89, b95, b951, b963, b987 = error_m(multmin, multmax, b, str(r'C:\Users\mcjara\OneDrive - Universidad Loyola Andalucía\Documentos\PycharmProjects\PSO_ASVs\Pruebas\Ori\Error'+str(seed[i])+'.xlsx'),
                      b20, b3, b12, b123, b145, b147, b156, b158, b159, b258, b321, b357, b369, b456, b4875, b541, b57, b65, b654, b741, b753, b756, b789, b8462, b852, b89, b95, b951, b963, b987)
        if multmax == 400:
            for k in range(len(mse)):
                mse400.append(mse[k])
        elif multmax == 800:
            for k in range(len(mse)):
                mse800.append(mse[k])
        elif multmax == 1200:
            for k in range(len(mse)):
                mse1200.append(mse[k])
        elif multmax == 1600:
            for k in range(len(mse)):
                mse1600.append(mse[k])
        elif multmax == 2000:
            for k in range(len(mse)):
                mse2000.append(mse[k])
        elif multmax == 2400:
            for k in range(len(mse)):
                mse2400.append(mse[k])
        elif multmax == 2800:
            for k in range(len(mse)):
                mse2800.append(mse[k])
        elif multmax == 3200:
            for k in range(len(mse)):
                mse3200.append(mse[k])
        elif multmax == 3600:
            for k in range(len(mse)):
                mse3600.append(mse[k])
        elif multmax == 4000:
            for k in range(len(mse)):
                mse4000.append(mse[k])
        elif multmax == 4400:
            for k in range(len(mse)):
                mse4400.append(mse[k])
        elif multmax == 4800:
            for k in range(len(mse)):
                mse4800.append(mse[k])
        elif multmax == 5200:
            for k in range(len(mse)):
                mse5200.append(mse[k])
        elif multmax == 5600:
            for k in range(len(mse)):
                mse5600.append(mse[k])
        elif multmax == 6000:
            for k in range(len(mse)):
                mse6000.append(mse[k])
    multmin = multmax
    print(multmax)
    print('in')

mse_mean = list()
mse400 = np.array(mse400)
mse800 = np.array(mse800)
mse1200 = np.array(mse1200)
mse1600 = np.array(mse1600)
mse2000 = np.array(mse2000)
mse2400 = np.array(mse2400)
mse2800 = np.array(mse2800)
mse3200 = np.array(mse3200)
mse3600 = np.array(mse3600)
mse4000 = np.array(mse4000)
mse4400 = np.array(mse4400)
mse4800 = np.array(mse4800)
mse5200 = np.array(mse5200)
mse5600 = np.array(mse5600)
mse6000 = np.array(mse6000)

mse_mean400 = np.mean(mse400)
mse_mean800 = np.mean(mse800)
mse_mean1200 = np.mean(mse1200)
mse_mean1600 = np.mean(mse1600)
mse_mean2000 = np.mean(mse2000)
mse_mean2400 = np.mean(mse2400)
mse_mean2800 = np.mean(mse2800)
mse_mean3200 = np.mean(mse3200)
mse_mean3600 = np.mean(mse3600)
mse_mean4000 = np.mean(mse4000)
mse_mean4400 = np.mean(mse4400)
mse_mean4800 = np.mean(mse4800)
mse_mean5200 = np.mean(mse5200)
mse_mean5600 = np.mean(mse5600)
mse_mean6000 = np.mean(mse6000)

mse_std400 = np.std(mse400)
mse_std800 = np.std(mse800)
mse_std1200 = np.std(mse1200)
mse_std1600 = np.std(mse1600)
mse_std2000 = np.std(mse2000)
mse_std2400 = np.std(mse2400)
mse_std2800 = np.std(mse2800)
mse_std3200 = np.std(mse3200)
mse_std3600 = np.std(mse3600)
mse_std4000 = np.std(mse4000)
mse_std4400 = np.std(mse4400)
mse_std4800 = np.std(mse4800)
mse_std5200 = np.std(mse5200)
mse_std5600 = np.std(mse5600)
mse_std6000 = np.std(mse6000)

mse_mean.append(mse_mean400)
mse_mean.append(mse_mean800)
mse_mean.append(mse_mean1200)
mse_mean.append(mse_mean1600)
mse_mean.append(mse_mean2000)
mse_mean.append(mse_mean2400)
mse_mean.append(mse_mean2800)
mse_mean.append(mse_mean3200)
mse_mean.append(mse_mean3600)
mse_mean.append(mse_mean4000)
mse_mean.append(mse_mean4400)
mse_mean.append(mse_mean4800)
mse_mean.append(mse_mean5200)
mse_mean.append(mse_mean5600)
mse_mean.append(mse_mean6000)

mse_std.append(mse_std400)
mse_std.append(mse_std800)
mse_std.append(mse_std1200)
mse_std.append(mse_std1600)
mse_std.append(mse_std2000)
mse_std.append(mse_std2400)
mse_std.append(mse_std2800)
mse_std.append(mse_std3200)
mse_std.append(mse_std3600)
mse_std.append(mse_std4000)
mse_std.append(mse_std4400)
mse_std.append(mse_std4800)
mse_std.append(mse_std5200)
mse_std.append(mse_std5600)
mse_std.append(mse_std6000)




wb10 = openpyxl.Workbook()
hoja10 = wb10.active
hoja10.append(mse_mean)
wb10.save(str(r'C:\Users\mcjara\OneDrive - Universidad Loyola Andalucía\Documentos\PycharmProjects\PSO_ASVs\Pruebas\Ori\MSE_Mean.xlsx'))
wb20 = openpyxl.Workbook()
hoja20 = wb20.active
hoja20.append(mse_std)
wb20.save(str(r'C:\Users\mcjara\OneDrive - Universidad Loyola Andalucía\Documentos\PycharmProjects\PSO_ASVs\Pruebas\Ori\MSE_Std.xlsx'))
wb30 = openpyxl.Workbook()
hoja30 = wb30.active
hoja30.append(mult)
wb30.save(str(r'C:\Users\mcjara\OneDrive - Universidad Loyola Andalucía\Documentos\PycharmProjects\PSO_ASVs\Pruebas\Ori\MSE_Mult.xlsx'))