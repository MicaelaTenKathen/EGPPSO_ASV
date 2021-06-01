import openpyxl
import numpy as np
import pandas as pd


def savexlsxori(MSE_data, sigma_data, mu_data, distances, w1, w2, w3, w4):
    for i in range(len(mu_data)):
        mu_data[i] = float(mu_data[i])
    #xcl = {'Seed': data[0],
     #      'GEN': data[1],
      #     'Time': data[2],
       #    'MSE_GEN': MSE_data[-1],
        #   'Avr_dist': np.mean(distances)}
    #xcl2 ={'Error': MSE_data,
     #      'Mean': mu_data,
      #     'Sigma': sigma_data}
    #df = pd.DataFrame(xcl, columns=['Seed', 'GEN', 'Time', 'MSE_GEN', 'Avr_dist'], index=[1])
    #df.to_excel(e5, 'Data')

    wb1 = openpyxl.Workbook()
    hoja1 = wb1.active
    hoja1.append(MSE_data)
    wb1.save(w1)
    wb2 = openpyxl.Workbook()
    hoja2 = wb2.active
    hoja2.append(sigma_data)
    wb2.save(w2)
    wb3 = openpyxl.Workbook()
    hoja3 = wb3.active
    hoja3.append(mu_data)
    wb3.save(w3)
    wb4 = openpyxl.Workbook()
    hoja4 = wb4.active
    hoja4.append(list(distances))
    wb4.save(w4)

