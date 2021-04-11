import openpyxl
import numpy as np
import matplotlib.pyplot as plt

def mean_std(n10, n15, n18, n20, n22, n25, n27, n30, n34, n35, n37, n40, n42, n45, n47, n52, n57, n62, n64, n67):
    q = 0
    psogp10 = []
    psogp15 = []
    psogp18 = []
    psogp20 = []
    psogp22 = []
    psogp25 = []
    psogp27 = []
    psogp30 = []
    psogp34 = []
    psogp35 = []
    psogp37 = []
    psogp40 = []
    psogp42 = []
    psogp45 = []
    psogp47 = []
    psogp52 = []
    psogp57 = []
    psogp62 = []
    psogp64 = []
    psogp67 = []

    wb1 = openpyxl.load_workbook(n10)
    hoja1 = wb1.active
    wb2 = openpyxl.load_workbook(n15)
    hoja2 = wb2.active
    wb3 = openpyxl.load_workbook(n18)
    hoja3 = wb3.active
    wb4 = openpyxl.load_workbook(n20)
    hoja4 = wb4.active
    wb5 = openpyxl.load_workbook(n22)
    hoja5 = wb5.active
    wb6 = openpyxl.load_workbook(n25)
    hoja6 = wb6.active
    wb7 = openpyxl.load_workbook(n27)
    hoja7 = wb7.active
    wb8 = openpyxl.load_workbook(n30)
    hoja8 = wb8.active
    wb9 = openpyxl.load_workbook(n34)
    hoja9 = wb9.active
    wb10 = openpyxl.load_workbook(n35)
    hoja10 = wb10.active
    wb11 = openpyxl.load_workbook(n37)
    hoja11 = wb11.active
    wb12 = openpyxl.load_workbook(n40)
    hoja12 = wb12.active
    wb13 = openpyxl.load_workbook(n42)
    hoja13 = wb13.active
    wb14 = openpyxl.load_workbook(n45)
    hoja14 = wb14.active
    wb15 = openpyxl.load_workbook(n47)
    hoja15 = wb15.active
    wb16 = openpyxl.load_workbook(n52)
    hoja16 = wb16.active
    wb17 = openpyxl.load_workbook(n57)
    hoja17 = wb17.active
    wb18 = openpyxl.load_workbook(n62)
    hoja18 = wb18.active
    wb19 = openpyxl.load_workbook(n64)
    hoja19 = wb19.active
    wb20 = openpyxl.load_workbook(n67)
    hoja20 = wb20.active

    while q < 201:
        q += 1
        cel1 = hoja1.cell(row=1, column=q)
        psogp10.append(cel1.value)
        cel2 = hoja2.cell(row=1, column=q)
        psogp15.append(cel2.value)
        cel3 = hoja3.cell(row=1, column=q)
        psogp18.append(cel3.value)
        cel4 = hoja4.cell(row=1, column=q)
        psogp20.append(cel4.value)
        cel5 = hoja5.cell(row=1, column=q)
        psogp22.append(cel5.value)
        cel6 = hoja6.cell(row=1, column=q)
        psogp25.append(cel6.value)
        cel7 = hoja7.cell(row=1, column=q)
        psogp27.append(cel7.value)
        cel8 = hoja8.cell(row=1, column=q)
        psogp30.append(cel8.value)
        cel9 = hoja9.cell(row=1, column=q)
        psogp34.append(cel9.value)
        cel10 = hoja10.cell(row=1, column=q)
        psogp35.append(cel10.value)
        cel11 = hoja11.cell(row=1, column=q)
        psogp37.append(cel11.value)
        cel12 = hoja12.cell(row=1, column=q)
        psogp40.append(cel12.value)
        cel13 = hoja13.cell(row=1, column=q)
        psogp42.append(cel13.value)
        cel14 = hoja14.cell(row=1, column=q)
        psogp45.append(cel14.value)
        cel15 = hoja15.cell(row=1, column=q)
        psogp47.append(cel15.value)
        cel16 = hoja16.cell(row=1, column=q)
        psogp52.append(cel16.value)
        cel17 = hoja17.cell(row=1, column=q)
        psogp57.append(cel17.value)
        cel18 = hoja18.cell(row=1, column=q)
        psogp62.append(cel18.value)
        cel19 = hoja19.cell(row=1, column=q)
        psogp64.append(cel19.value)
        cel20 = hoja20.cell(row=1, column=q)
        psogp67.append(cel20.value)

    p = q - 1
    n = 0
    mean_array = []
    std_array = []
    std196 = []
    mean = []
    it = []

    while n <= p:
        if n%20 == 0:
            mean.append(psogp10[n])
            mean.append(psogp15[n])
            mean.append(psogp18[n])
            mean.append(psogp20[n])
            mean.append(psogp22[n])
            mean.append(psogp25[n])
            mean.append(psogp27[n])
            mean.append(psogp30[n])
            mean.append(psogp34[n])
            mean.append(psogp35[n])
            mean.append(psogp37[n])
            mean.append(psogp40[n])
            mean.append(psogp42[n])
            mean.append(psogp45[n])
            mean.append(psogp47[n])
            mean.append(psogp52[n])
            mean.append(psogp57[n])
            mean.append(psogp62[n])
            mean.append(psogp64[n])
            mean.append(psogp67[n])
            meanarray = np.array(mean)
            mean_value = np.mean(meanarray)
            std_value = np.std(meanarray)
            stdvalue = 1.96 * std_value
            std196.append(stdvalue)
            mean_array.append(mean_value)
            std_array.append(std_value)
            mean = []
            it.append(n)
        n += 1
    it_array = np.array(it)
    return mean_array, std_array, std196, it, it_array

def comparison():
    width = 4
    y = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10]

    fig, ax = plt.subplots()
    #rects1 = ax.bar(it_array, meanpso_array1, width, color='b', yerr=stdpso_array1, label='UN + MU(1), PSO', alpha=0.9)
    rects2 = ax.bar(it_array + width, meanpso_array2, width, color='y', yerr=stdpso_array2, label='UN + MU(2), PSO', alpha=0.9)
    #rects3 = ax.bar(it_array + 2 * width, meanpso_array3, width, color='g', yerr=stdpso_array3, label='UN + MU(3), PSO', alpha=0.9)
    rects4 = ax.bar(it_array + 3 * width, meanpsogp_array, width, color='c', yerr=stdpsogp_array, label='GP-based PSO',
                    alpha=0.9)

    # Add some text for labels, title and custom x-axis tick labels, etc.
    ax.set_ylabel('MSE', fontsize=12)
    ax.set_xlabel('Iterations', fontsize=12)
    ax.set_xticks(it)
    ax.set_yticks(y)
    ax.set_ylim([0, 7])
    ax.set_yticklabels(y, fontsize=12)
    ax.set_xticklabels(it, fontsize=12)
    ax.legend(loc=2, fontsize=12)
    ax.grid(True)

#    def autolabel(rects):
 #       """Attach a text label above each bar in *rects*, displaying its height."""
  #      for rect in rects:
   #         height = rect.get_height()
    #        ax.annotate('{}'.format(height),
     #                   xy=(rect.get_x() + rect.get_width() / 2, height),
      #                  xytext=(0, 3),  # 3 points vertical offset
       #                 textcoords="offset points",
        #                ha='center', va='bottom')

#    autolabel(rects1)
 #   autolabel(rects2)

    fig.tight_layout()

    plt.show()

n10 = str('Pruebas/PSOError10.xlsx')
n15 = str('Pruebas/PSOError15.xlsx')
n18 = str('Pruebas/PSOError18.xlsx')
n20 = str('Pruebas/PSOError20.xlsx')
n22 = str('Pruebas/PSOError22.xlsx')
n25 = str('Pruebas/PSOError25.xlsx')
n27 = str('Pruebas/PSOError27.xlsx')
n30 = str('Pruebas/PSOError30.xlsx')
n34 = str('Pruebas/PSOError34.xlsx')
n35 = str('Pruebas/PSOError35.xlsx')
n37 = str('Pruebas/PSOError37.xlsx')
n40 = str('Pruebas/PSOError40.xlsx')
n42 = str('Pruebas/PSOError42.xlsx')
n45 = str('Pruebas/PSOError45.xlsx')
n47 = str('Pruebas/PSOError47.xlsx')
n52 = str('Pruebas/PSOError52.xlsx')
n57 = str('Pruebas/PSOError57.xlsx')
n62 = str('Pruebas/PSOError62.xlsx')
n64 = str('Pruebas/PSOError64.xlsx')
n67 = str('Pruebas/PSOError67.xlsx')
meanpso_array2, stdpso_array2, stdpso1962, it, it_array = mean_std(n10, n15, n18, n20, n22, n25, n27, n30, n34, n35, n37,
                                                                n40, n42, n45, n47, n52, n57, n62, n64, n67)
print(meanpso_array2, stdpso_array2, stdpso1962)

n10 = str('Pruebas/Error2201130610.xlsx')
n15 = str('Pruebas/Error2201130615.xlsx')
n18 = str('Pruebas/Error2201130618.xlsx')
n20 = str('Pruebas/Error2201130620.xlsx')
n22 = str('Pruebas/Error2201130622.xlsx')
n25 = str('Pruebas/Error2201130625.xlsx')
n27 = str('Pruebas/Error2201130627.xlsx')
n30 = str('Pruebas/Error2201130630.xlsx')
n34 = str('Pruebas/Error2201130634.xlsx')
n35 = str('Pruebas/Error2201130635.xlsx')
n37 = str('Pruebas/Error2201130637.xlsx')
n40 = str('Pruebas/Error2201130640.xlsx')
n42 = str('Pruebas/Error2201130642.xlsx')
n45 = str('Pruebas/Error2201130645.xlsx')
n47 = str('Pruebas/Error2201130647.xlsx')
n52 = str('Pruebas/Error2201130652.xlsx')
n57 = str('Pruebas/Error2201130657.xlsx')
n62 = str('Pruebas/Error2201130662.xlsx')
n64 = str('Pruebas/Error2201130664.xlsx')
n67 = str('Pruebas/Error2201130667.xlsx')
meanpsogp_array, stdpsogp_array, stdpsogp196, it, it_array = mean_std(n10, n15, n18, n20, n22, n25, n27, n30, n34, n35,
                                                                      n37, n40, n42, n45, n47, n52, n57, n62, n64, n67)
print(meanpsogp_array, stdpsogp_array, stdpsogp196)

# n10 = str('ErrorMU101.xlsx')
# n15 = str('ErrorMU151.xlsx')
# n18 = str('ErrorMU181.xlsx')
# n20 = str('ErrorMU201.xlsx')
# n22 = str('ErrorMU221.xlsx')
# n25 = str('ErrorMU251.xlsx')
# n27 = str('ErrorMU271.xlsx')
# n30 = str('ErrorMU301.xlsx')
# n34 = str('ErrorMU341.xlsx')
# n35 = str('ErrorMU351.xlsx')
# n37 = str('ErrorMU371.xlsx')
# n40 = str('ErrorMU401.xlsx')
# n42 = str('ErrorMU421.xlsx')
# n45 = str('ErrorMU451.xlsx')
# n47 = str('ErrorMU471.xlsx')
# n52 = str('ErrorMU521.xlsx')
# n57 = str('ErrorMU571.xlsx')
# n62 = str('ErrorMU621.xlsx')
# n64 = str('ErrorMU641.xlsx')
# n67 = str('ErrorMU671.xlsx')
# meanpso_array1, stdpso_array1, stdpso1961, it, it_array = mean_std(n10, n15, n18, n20, n22, n25, n27, n30, n34, n35, n37,
#                                                                 n40, n42, n45, n47, n52, n57, n62, n64, n67)
# print(meanpso_array1, stdpso_array1, stdpso1961)
#
# n10 = str('ErrorMU103.xlsx')
# n15 = str('ErrorMU153.xlsx')
# n18 = str('ErrorMU183.xlsx')
# n20 = str('ErrorMU203.xlsx')
# n22 = str('ErrorMU223.xlsx')
# n25 = str('ErrorMU253.xlsx')
# n27 = str('ErrorMU273.xlsx')
# n30 = str('ErrorMU303.xlsx')
# n34 = str('ErrorMU343.xlsx')
# n35 = str('ErrorMU353.xlsx')
# n37 = str('ErrorMU373.xlsx')
# n40 = str('ErrorMU403.xlsx')
# n42 = str('ErrorMU423.xlsx')
# n45 = str('ErrorMU453.xlsx')
# n47 = str('ErrorMU473.xlsx')
# n52 = str('ErrorMU523.xlsx')
# n57 = str('ErrorMU573.xlsx')
# n62 = str('ErrorMU623.xlsx')
# n64 = str('ErrorMU643.xlsx')
# n67 = str('ErrorMU673.xlsx')
# meanpso_array3, stdpso_array3, stdpso1963, it, it_array = mean_std(n10, n15, n18, n20, n22, n25, n27, n30, n34, n35,
#                                                                       n37, n40, n42, n45, n47, n52, n57, n62, n64, n67)
# print(meanpso_array3, stdpso_array3, stdpso1963)

comparison()

#fig, ax = plt.subplots()
#ax.bar(it, meanpso_array,
 #      yerr=stdpso_array,
  #     width=10,
   #    color='blue',
    #   align='center',
     #  alpha=0.5,
      # ecolor='black',
       #capsize=5)
#ax.set_ylabel('MSE')
#ax.set_xticks(it)
#ax.set_title('PSOGP Mean Square Error')
#ax.set_xlabel('Iterations')
#ax.yaxis.grid('True')

#plt.tight_layout()
#plt.show()