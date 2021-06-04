import openpyxl
import numpy as np
import matplotlib.pyplot as plt


def extract(n10, n15, n18, n20, n22, n25, n27, n30, n34):
    q = 0

    meanori = []
    stdori = []
    meangp = []
    stdgp = []
    meannew = []
    stdnew = []
    mult = []
    meancase = []
    stdcase = []

    wb1 = openpyxl.load_workbook(n10)
    hoja1 = wb1.active
    wb2 = openpyxl.load_workbook(n15)
    hoja2 = wb2.active
    wb3 = openpyxl.load_workbook(n18)
    hoja3 = wb3.active
    wb4 = openpyxl.load_workbook(n20)
    hoja4 = wb4.active
    wb5 = openpyxl.load_workbook(n22)
    hoja5 = wb5.active
    wb6 = openpyxl.load_workbook(n25)
    hoja6 = wb6.active
    wb7 = openpyxl.load_workbook(n27)
    hoja7 = wb7.active
    wb8 = openpyxl.load_workbook(n30)
    hoja8 = wb8.active
    wb9 = openpyxl.load_workbook(n34)
    hoja9 = wb9.active

    while q < 15:
        q += 1
        cel1 = hoja1.cell(row=1, column=q)
        meanori.append(cel1.value)
        cel2 = hoja2.cell(row=1, column=q)
        stdori.append(cel2.value)
        cel3 = hoja3.cell(row=1, column=q)
        meangp.append(cel3.value)
        cel4 = hoja4.cell(row=1, column=q)
        stdgp.append(cel4.value)
        cel5 = hoja5.cell(row=1, column=q)
        meannew.append(cel5.value)
        cel6 = hoja6.cell(row=1, column=q)
        stdnew.append(cel6.value)
        cel7 = hoja7.cell(row=1, column=q)
        mult.append(cel7.value)
        cel8 = hoja8.cell(row=1, column=q)
        meancase.append(cel8.value)
        cel9 = hoja9.cell(row=1, column=q)
        stdcase.append(cel9.value)

    return meanori, stdori, meangp, stdgp, meannew, stdnew, mult, meancase, stdcase


def comparison(meanori, stdori, meangp, stdgp, meannew, stdnew, mult, meancase, stdcase):
    width = 80

    fig, ax = plt.subplots()
    ax.bar(mult - 1 * width, meanori, width, color='b', yerr=stdori, label='Original PSO', alpha=0.9)
    ax.bar(mult, meangp, width, color='y', yerr=stdgp, label='GP-based PSO', alpha=0.9)
    ax.bar(mult + 1 * width, meannew, width, color='c', yerr=stdnew, label='Enhanced GP-based PSO (Case 0)', alpha=0.9)
    ax.bar(mult + 2 * width, meancase, width, color='g', yerr=stdcase, label='Enhanced GP-based PSO (Case All)',
           alpha=0.9)

    ax.set_ylabel('MSE', fontsize=12)
    ax.set_xlabel('Iterations', fontsize=12)

    ax.legend(loc=1, fontsize=12)
    ax.grid(True)


n10 = str(r'C:\Users\mcjara\OneDrive - Universidad Loyola '
          r'Andalucía\Documentos\PycharmProjects\EGPPSO_ASV\Pruebas\Ori\MSE_Mean.xlsx')
n15 = str(r'C:\Users\mcjara\OneDrive - Universidad Loyola '
          r'Andalucía\Documentos\PycharmProjects\EGPPSO_ASV\Pruebas\Ori\MSE_Std.xlsx')
n18 = str(r'C:\Users\mcjara\OneDrive - Universidad Loyola '
          r'Andalucía\Documentos\PycharmProjects\EGPPSO_ASV\Pruebas\GP\MSE_Mean.xlsx')
n20 = str(r'C:\Users\mcjara\OneDrive - Universidad Loyola '
          r'Andalucía\Documentos\PycharmProjects\EGPPSO_ASV\Pruebas\GP\MSE_Std.xlsx')
n22 = str(r'C:\Users\mcjara\OneDrive - Universidad Loyola '
          r'Andalucía\Documentos\PycharmProjects\EGPPSO_ASV\Pruebas\NewGP\MSE_Mean.xlsx')
n25 = str(r'C:\Users\mcjara\OneDrive - Universidad Loyola '
          r'Andalucía\Documentos\PycharmProjects\EGPPSO_ASV\Pruebas\NewGP\MSE_Std.xlsx')
n27 = str(r'C:\Users\mcjara\OneDrive - Universidad Loyola '
          r'Andalucía\Documentos\PycharmProjects\EGPPSO_ASV\Pruebas\Ori\MSE_Mult.xlsx')
n30 = str(r'C:\Users\mcjara\OneDrive - Universidad Loyola '
          r'Andalucía\Documentos\PycharmProjects\EGPPSO_ASV\Pruebas\NewGPCaseA\MSE_Mean.xlsx')
n34 = str(r'C:\Users\mcjara\OneDrive - Universidad Loyola '
          r'Andalucía\Documentos\PycharmProjects\EGPPSO_ASV\Pruebas\NewGPCaseA\MSE_Std.xlsx')


meanori, stdori, meangp, stdgp, meannew, stdnew, mult, meancase, stdcase = extract(n10, n15, n18, n20, n22, n25, n27,
                                                                                   n30, n34)

mult_array = np.array(mult)

mult_array1 = np.array([400, 800, 1200, 1600, 2000, 2400, 2800, 3200, 3600, 4000, 4400, 4800, 5200, 5600, 6000])

comparison(meanori, stdori, meangp, stdgp, meannew, stdnew, mult_array1, meancase, stdcase)
